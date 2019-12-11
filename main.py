#!/usr/bin/python
import sys
import re
import csv
import openpyxl

variables = [
    'AX', 'SI', 'CS', 'Stack +0',
    'BX', 'DI', 'DS', 'Stack +2',
    'CX', 'BP', 'ES', 'Stack +4',
    'DX', 'SP', 'SS', 'Stack +6',
    'OF', 'DF', 'IF', 'SF', 'ZF', 'AF', 'PF', 'CF',
    'current_address', 'command'
]
equals = {
    'tfp' : 'trace_file_path',
    'tbfp' : 'table_file_path',
    'sn'  : 'sheet_name',
    'tt'  : 'tab_top',
}

class Iteration:
    def __init__(self, reg_values, flags_values, current_address, command, address):
        values = reg_values+flags_values+current_address+command
        self.values = dict.fromkeys(variables)
        for index, key in enumerate(self.values.keys()):
            self.values[key] = values[index]
        if len(address) == 0:
            self.values['address'] = ''
        else:
            result = 0
            expr = address[0][1:-1]
            opers = re.findall('[+-]', expr)
            vals = re.split('[+-]', expr)
            for i, val in enumerate(vals):
                if val in variables:
                    # print('true')
                    vals[i] = int(eval('0x' + self.get(val)))
                else:
                    vals[i] = int(eval('0x' + val))

            for i, oper in enumerate(opers):
                result += eval(str(vals[i]) + oper + str(vals[i + 1]))

            self.values['address'] = (hex(result))[2:]
            print(self.get('address'))

    def get(self, key):
        return self.values[key]

def parse_args(args):
    result = dict()
    for arg in args:
        key, value = arg.split('=')
        if key in equals.keys():
            key = equals[key]
        result[key] = value
    return result

def to_excel(excel_file_path, sheet_name, tab_top, iterations, mask):
    try:
        excel_file = openpyxl.load_workbook(filename = excel_file_path)
    except:
        excel_file = openpyxl.Workbook()

    if sheet_name not in excel_file.sheetnames:
        excel_file.create_sheet(sheet_name)
    sheet = excel_file[sheet_name]

    tab_top = int(tab_top)

    for i in range(len(iterations)):
        for k, arg in enumerate(mask):
            sheet.cell(row=tab_top+i+1, column=k+1).value = iterations[i].get(mask[k])

    excel_file.save(excel_file_path)

def to_csv(csv_file_path, iterations, mask):
    csv_file = open(csv_file_path, 'w')
    writer = csv.DictWriter(csv_file, fieldnames=mask, quoting=csv.QUOTE_ALL)
    writer.writeheader()

    for iteration in iterations:
        writer.writerow({key: iteration.get(key) for key in mask})

    csv_file.close()

def parse_afdpro(
    trace_file_path='trace.txt', 
    table_file_path='trace.csv',
    sheet_name='trace',
    tab_top=0,
    type='csv',
    mask=['current_address', 'command', 'AX', 'BX', 'CX', 'DX', 'OF', 'DF', 'IF', 'SF', 'ZF', 'AF', 'PF', 'CF', 'Stack +0', 'address']
):

    with open(trace_file_path, 'r') as file:
        lines = file.readlines()[3:-2]
    commands = [''.join(lines[i:i + 4]) for i in range(0, len(lines), 4)]

    iterations = []

    for i, command in enumerate(commands):
        temp = re.findall('\w+', command[:12])
        # print(re.search(r'\[[A-Z0-9+-]+\]', command[12:41]), command[12:41])
        iterations.append( Iteration(
            reg_values      = re.findall('[0-9A-F]{4}', command[42:]), # hex number ignore first 42 chars
            flags_values    = re.findall('[0-1]', command[-63:-40]), # 0 or 1
            current_address = temp[:1],
            command         = temp[1:],
            address         = re.findall(r'\[[A-Z0-9+-]+\]', command[12:41])
        ))

    if type == 'csv':
        to_csv(table_file_path, iterations, mask)
    elif type == 'excel':
        pass
        to_excel(table_file_path, sheet_name, tab_top, iterations, mask)
    file.close()

if __name__ == '__main__':
    parse_afdpro(**parse_args(sys.argv[1:]))